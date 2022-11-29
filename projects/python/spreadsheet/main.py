import openpyxl

""" 
Will calculate a serie of operations within the passed inventory.xlsx
and will generate a new generated_inventory.xlsx with extra information
"""
inventory_file = openpyxl.load_workbook("inventory.xlsx")
sheet = inventory_file["Sheet 1"]
HEADERS_ROW = 2
FIRST_ROW = 3

ID_COLUMN = 1
INVENTORY_COLUMN = 2
PRICE_COLUMN = 3
MANUFACTURER_COLUMN = 4
NEW_INVENTORY_PRICE_COLUMN = 5

products_per_manufacturer = {}
value_per_manufacturer = {}
products_under_10_items = {}


def round_2(number):
    """Round the passed float to the closest 2 decimals value"""
    print(f"original:{number}, rounded:{round(number, 2)}")
    return round(number, 2)


for row in range(FIRST_ROW, sheet.max_row + 1):
    product_id = sheet.cell(row, ID_COLUMN).value
    inventory = sheet.cell(row, INVENTORY_COLUMN).value
    price = sheet.cell(row, PRICE_COLUMN).value
    manufacturer = sheet.cell(row, MANUFACTURER_COLUMN).value

    # 1 Calculate total products per manufacturer
    if manufacturer in products_per_manufacturer:
        products_per_manufacturer[manufacturer] = products_per_manufacturer.get(manufacturer) + 1
    else:
        products_per_manufacturer[manufacturer] = 1

    # 2 Calculate total value of inventory per manufacturer
    if manufacturer in value_per_manufacturer:
        value_per_manufacturer[manufacturer] = round_2(value_per_manufacturer.get(manufacturer) + (inventory * price))
    else:
        value_per_manufacturer[manufacturer] = round_2(inventory * price)

    # 3 Find products with inventories smaller than 10 items
    if inventory < 10:
        products_under_10_items[product_id] = inventory

    # 4.1 Add new column with total inventory price
    inventory_price = sheet.cell(row, NEW_INVENTORY_PRICE_COLUMN)
    inventory_price.value = round_2(inventory * price)

# 4.2 Create header for the new column and save file
sheet.cell(HEADERS_ROW, NEW_INVENTORY_PRICE_COLUMN).value = "Inventory Price"
inventory_file.save("generated_inventory.xlsx")

print(f"Total products per manufacturer:\n{products_per_manufacturer}\n")
print(f"Total values per manufacturer:\n{value_per_manufacturer}\n")
print(f"Products with inventories smaller than 10 items:\n{products_under_10_items}\n")
print("generated_inventory.xlsx was generated")
